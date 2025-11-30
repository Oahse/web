from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from sqlalchemy.orm import aliased
from typing import List, Tuple
from datetime import datetime, timedelta
from models.user import User
from models.product import Product, ProductVariant, Category
from models.order import Order, OrderItem
from models.review import Review
import csv
import io


class AnalyticsService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_dashboard_data(self, user_id: str, user_role: str, start_date: datetime, end_date: datetime = None, filters: dict = None) -> dict:
        """Get dashboard analytics data based on user role with filters."""
        if end_date is None:
            end_date = datetime.now()
        if filters is None:
            filters = {}
            
        if user_role in ["Admin", "SuperAdmin"]:
            return await self._get_admin_dashboard_data(start_date, end_date, filters)
        elif user_role == "Supplier":
            return await self._get_supplier_dashboard_data(user_id, start_date, end_date, filters)
        else:
            return await self._get_customer_dashboard_data(user_id, start_date, end_date, filters)

    async def _get_admin_dashboard_data(self, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get admin dashboard data with filters."""
        # Build base conditions
        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        
        # Apply filters to queries
        filter_conditions = []
        if filters.get('order_status'):
            filter_conditions.append(Order.status == filters['order_status'])
        
        # For category and product filters, we need to join with OrderItem and Product
        needs_product_join = filters.get('category') or filters.get('product')
        
        # Total sales
        if needs_product_join:
            sales_query = select(func.sum(OrderItem.total_price)).join(
                Order, OrderItem.order_id == Order.id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if filters.get('category'):
                # Join with Category table to filter by category name
                sales_query = sales_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                sales_query = sales_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
        else:
            sales_query = select(func.sum(Order.total_amount)).where(date_conditions)
            if filter_conditions:
                sales_query = sales_query.where(and_(*filter_conditions))
                
        sales_result = await self.db.execute(sales_query)
        total_sales = sales_result.scalar() or 0

        # Total orders
        orders_query = select(func.count(Order.id.distinct())).where(date_conditions)
        if filter_conditions:
            orders_query = orders_query.where(and_(*filter_conditions))
        if needs_product_join:
            orders_query = select(func.count(Order.id.distinct())).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                orders_query = orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                orders_query = orders_query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filter_conditions:
                orders_query = orders_query.where(and_(*filter_conditions))
                
        orders_result = await self.db.execute(orders_query)
        total_orders = orders_result.scalar() or 0

        # Total users (with user segment filter)
        user_date_conditions = and_(
            User.created_at >= start_date,
            User.created_at <= end_date
        )
        users_query = select(func.count(User.id)).where(user_date_conditions)
        
        # Apply user segment filter
        if filters.get('user_segment'):
            if filters['user_segment'] == 'new':
                # Users with no orders
                users_query = select(func.count(User.id)).where(
                    and_(
                        user_date_conditions,
                        ~User.id.in_(select(Order.user_id.distinct()))
                    )
                )
            elif filters['user_segment'] == 'returning':
                # Users with more than one order
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.count(Order.id) > 1
                )
            elif filters['user_segment'] == 'vip':
                # Users with total spending > threshold (e.g., $1000)
                users_query = select(func.count(User.id.distinct())).join(
                    Order, User.id == Order.user_id
                ).where(user_date_conditions).group_by(User.id).having(
                    func.sum(Order.total_amount) > 1000
                )
                
        users_result = await self.db.execute(users_query)
        total_users = users_result.scalar() or 0

        # Total products
        products_query = select(func.count(Product.id))
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        total_products = products_result.scalar() or 0

        # Get sales trend and top products with filters
        sales_trend = await self.get_sales_trend(
            user_id="", 
            user_role="Admin", 
            days=(end_date - start_date).days,
            filters=filters
        )
        top_products = await self.get_top_products(
            user_id="", 
            user_role="Admin", 
            limit=5,
            filters=filters
        )

        # Order status distribution
        order_status_query = select(
            Order.status,
            func.count(Order.id).label("count")
        ).where(date_conditions)
        
        if needs_product_join:
            order_status_query = select(
                Order.status,
                func.count(Order.id.distinct()).label("count")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            if filters.get('category'):
                order_status_query = order_status_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                order_status_query = order_status_query.where(Product.name.ilike(f"%{filters['product']}%"))
        
        if filter_conditions:
            order_status_query = order_status_query.where(and_(*filter_conditions))
            
        order_status_query = order_status_query.group_by(Order.status)
        order_status_result = await self.db.execute(order_status_query)
        order_status_distribution = {
            row.status: row.count for row in order_status_result.all()}

        # User growth
        user_growth_query = select(
            func.date(User.created_at).label("date"),
            func.count(User.id).label("new_users")
        ).where(user_date_conditions).group_by(
            func.date(User.created_at)
        ).order_by(
            func.date(User.created_at)
        )
        user_growth_result = await self.db.execute(user_growth_query)
        user_growth = [{
            "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
            "new_users": row.new_users
        } for row in user_growth_result.all()]

        # Conversion rate (simplified: unique purchasing users / total users)
        purchasing_users_query = select(func.count(Order.user_id.distinct())).where(date_conditions)
        if filter_conditions:
            purchasing_users_query = purchasing_users_query.where(and_(*filter_conditions))
        purchasing_users_result = await self.db.execute(purchasing_users_query)
        unique_purchasing_users = purchasing_users_result.scalar() or 0

        conversion_rate = (unique_purchasing_users /
                           total_users * 100) if total_users > 0 else 0.0

        return {
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "total_users": total_users,
            "total_products": total_products,
            "conversion_rate": round(conversion_rate, 2),
            "average_order_value": float(total_sales / total_orders) if total_orders > 0 else 0,
            "top_products": top_products,
            "sales_trend": sales_trend,
            "order_status_distribution": order_status_distribution,
            "user_growth": user_growth
        }

    async def _get_supplier_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get supplier dashboard data with filters."""
        # Get supplier's products
        products_query = select(Product).where(Product.supplier_id == user_id)
        if filters.get('category'):
            products_query = products_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        products_result = await self.db.execute(products_query)
        products = products_result.scalars().all()

        # Calculate total sales and orders for the supplier
        order_item_alias = aliased(OrderItem)
        product_variant_alias = aliased(ProductVariant)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        supplier_orders_query = select(Order).join(
            order_item_alias, Order.id == order_item_alias.order_id
        ).join(
            product_variant_alias, order_item_alias.variant_id == product_variant_alias.id
        ).join(
            Product, product_variant_alias.product_id == Product.id
        ).where(
            and_(
                Product.supplier_id == user_id,
                date_conditions
            )
        )
        
        if filters.get('category'):
            supplier_orders_query = supplier_orders_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('order_status'):
            supplier_orders_query = supplier_orders_query.where(Order.status == filters['order_status'])
            
        supplier_orders_query = supplier_orders_query.distinct()

        supplier_orders_result = await self.db.execute(supplier_orders_query)
        supplier_orders = supplier_orders_result.scalars().all()

        total_sales = sum(order.total_amount for order in supplier_orders)
        total_orders = len(supplier_orders)

        # Calculate average rating for supplier's products
        reviews_query = select(func.avg(Review.rating)).join(
            Product, Review.product_id == Product.id
        ).where(
            Product.supplier_id == user_id
        )
        if filters.get('category'):
            reviews_query = reviews_query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        average_rating_result = await self.db.execute(reviews_query)
        average_rating = average_rating_result.scalar() or 0.0

        return {
            "total_products": len(products),
            "total_sales": float(total_sales),
            "total_orders": total_orders,
            "average_rating": round(average_rating, 2),
            "products": [
                {
                    "id": str(p.id),
                    "name": p.name,
                    "rating": p.rating,
                    "review_count": p.review_count
                }
                for p in products[:5]
            ]
        }

    async def _get_customer_dashboard_data(self, user_id: str, start_date: datetime, end_date: datetime, filters: dict) -> dict:
        """Get customer dashboard data with filters."""
        # Get customer's orders
        date_conditions = and_(
            Order.user_id == user_id,
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )
        orders_query = select(Order).where(date_conditions)
        
        if filters.get('order_status'):
            orders_query = orders_query.where(Order.status == filters['order_status'])
            
        orders_result = await self.db.execute(orders_query)
        orders = orders_result.scalars().all()

        total_spent = sum(order.total_amount for order in orders)

        return {
            "total_orders": len(orders),
            "total_spent": float(total_spent),
            "average_order_value": float(total_spent / len(orders)) if orders else 0,
            "recent_orders": [
                {
                    "id": str(order.id),
                    "status": order.status,
                    "total_amount": order.total_amount,
                    "created_at": order.created_at.isoformat()
                }
                for order in orders[:5]
            ]
        }

    async def get_sales_trend(self, user_id: str, user_role: str, days: int = 30, filters: dict = None) -> List[dict]:
        """Get sales trend data grouped by date with filters."""
        if filters is None:
            filters = {}
            
        end_date = datetime.now()
        start_date = end_date - timedelta(days=days)

        date_conditions = and_(
            Order.created_at >= start_date,
            Order.created_at <= end_date
        )

        # Check if we need product joins for filtering
        needs_product_join = filters.get('category') or filters.get('product') or user_role == "Supplier"

        if needs_product_join:
            # Use OrderItem-based query for product filtering
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(OrderItem.total_price).label("sales"),
                func.count(Order.id.distinct()).label("orders")
            ).join(
                OrderItem, Order.id == OrderItem.order_id
            ).join(
                ProductVariant, OrderItem.variant_id == ProductVariant.id
            ).join(
                Product, ProductVariant.product_id == Product.id
            ).where(date_conditions)
            
            if user_role == "Supplier":
                query = query.where(Product.supplier_id == user_id)
            if filters.get('category'):
                query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
            if filters.get('product'):
                query = query.where(Product.name.ilike(f"%{filters['product']}%"))
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )
        else:
            # Use simple Order-based query
            query = select(
                func.date(Order.created_at).label("date"),
                func.sum(Order.total_amount).label("sales"),
                func.count(Order.id).label("orders")
            ).where(date_conditions)
            
            if filters.get('order_status'):
                query = query.where(Order.status == filters['order_status'])
                
            query = query.group_by(
                func.date(Order.created_at)
            ).order_by(
                func.date(Order.created_at)
            )

        result = await self.db.execute(query)
        trend_data = []
        for row in result.all():
            trend_data.append({
                "date": row.date.strftime("%Y-%m-%d") if hasattr(row.date, 'strftime') else str(row.date),
                "sales": float(row.sales or 0),
                "orders": int(row.orders or 0)
            })

        return trend_data

    async def get_top_products(self, user_id: str, user_role: str, limit: int = 10, filters: dict = None) -> List[dict]:
        """Get top performing products by revenue with filters."""
        if filters is None:
            filters = {}
            
        # Get the most recent orders to ensure fresh data
        from sqlalchemy.orm import selectinload
        
        query = select(
            Product.id,
            Product.name,
            func.sum(OrderItem.quantity).label("total_sales_quantity"),
            func.sum(OrderItem.total_price).label("total_revenue")
        ).join(
            ProductVariant, Product.id == ProductVariant.product_id
        ).join(
            OrderItem, ProductVariant.id == OrderItem.variant_id
        ).join(
            Order, OrderItem.order_id == Order.id
        ).where(
            Order.status.in_(['confirmed', 'shipped', 'delivered', 'processing'])
        )
        
        if user_role == "Supplier":
            query = query.where(Product.supplier_id == user_id)
        if filters.get('category'):
            query = query.join(Category, Product.category_id == Category.id).where(Category.name == filters['category'])
        if filters.get('product'):
            query = query.where(Product.name.ilike(f"%{filters['product']}%"))
        if filters.get('order_status'):
            query = query.where(Order.status == filters['order_status'])
            
        query = query.group_by(
            Product.id,
            Product.name
        ).order_by(
            func.sum(OrderItem.total_price).desc()
        ).limit(limit)

        result = await self.db.execute(query)
        products_data = []
        for row in result.all():
            # Get product image from first variant
            image_query = select(ProductVariant).where(
                ProductVariant.product_id == row.id
            ).options(selectinload(ProductVariant.images)).limit(1)
            variant_result = await self.db.execute(image_query)
            variant = variant_result.scalar_one_or_none()
            
            image_url = None
            if variant and variant.images:
                primary_image = next((img for img in variant.images if img.is_primary), None)
                image_url = primary_image.url if primary_image else (variant.images[0].url if variant.images else None)
            
            products_data.append({
                "id": str(row.id),
                "name": row.name,
                "sales": int(row.total_sales_quantity or 0),
                "revenue": float(row.total_revenue or 0),
                "image_url": image_url
            })

        return products_data

    async def export_data(self, data: dict, format: str, export_type: str) -> Tuple[bytes, str, str]:
        """
        Export analytics data to CSV or Excel format.
        
        Args:
            data: The analytics data dictionary
            format: 'csv' or 'xlsx'
            export_type: Type of export (e.g., 'dashboard')
            
        Returns:
            Tuple of (file_content, content_type, filename)
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if format == 'csv':
            return await self._export_csv(data, export_type, timestamp)
        elif format == 'xlsx':
            return await self._export_excel(data, export_type, timestamp)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    async def _export_csv(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write summary section
        writer.writerow(['Analytics Summary'])
        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Sales', f"${data.get('total_sales', 0):.2f}"])
        writer.writerow(['Total Orders', data.get('total_orders', 0)])
        writer.writerow(['Total Users', data.get('total_users', 0)])
        writer.writerow(['Total Products', data.get('total_products', 0)])
        writer.writerow(['Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"])
        writer.writerow(['Average Order Value', f"${data.get('average_order_value', 0):.2f}"])
        writer.writerow([])
        
        # Write sales trend section
        writer.writerow(['Sales Trend'])
        writer.writerow(['Date', 'Sales', 'Orders'])
        for item in data.get('sales_trend', []):
            writer.writerow([
                item.get('date', ''),
                f"${item.get('sales', 0):.2f}",
                item.get('orders', 0)
            ])
        writer.writerow([])
        
        # Write top products section
        writer.writerow(['Top Products'])
        writer.writerow(['Product Name', 'Units Sold', 'Revenue'])
        for product in data.get('top_products', []):
            writer.writerow([
                product.get('name', ''),
                product.get('sales', 0),
                f"${product.get('revenue', 0):.2f}"
            ])
        writer.writerow([])
        
        # Write order status distribution
        writer.writerow(['Order Status Distribution'])
        writer.writerow(['Status', 'Count'])
        for status, count in data.get('order_status_distribution', {}).items():
            writer.writerow([status, count])
        
        # Convert to bytes
        content = output.getvalue().encode('utf-8')
        filename = f"analytics_{export_type}_{timestamp}.csv"
        
        return content, 'text/csv', filename
    
    async def _export_excel(self, data: dict, export_type: str, timestamp: str) -> Tuple[bytes, str, str]:
        """Export data as Excel."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise ImportError("openpyxl is required for Excel export. Install it with: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Summary"
        
        # Define styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        
        # Write summary section
        row = 1
        ws.cell(row=row, column=1, value='Analytics Summary').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Metric').font = header_font
        ws.cell(row=row, column=2, value='Value').font = header_font
        row += 1
        
        summary_data = [
            ('Total Sales', f"${data.get('total_sales', 0):.2f}"),
            ('Total Orders', data.get('total_orders', 0)),
            ('Total Users', data.get('total_users', 0)),
            ('Total Products', data.get('total_products', 0)),
            ('Conversion Rate', f"{data.get('conversion_rate', 0):.2f}%"),
            ('Average Order Value', f"${data.get('average_order_value', 0):.2f}")
        ]
        
        for metric, value in summary_data:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        row += 2
        
        # Write sales trend section
        ws.cell(row=row, column=1, value='Sales Trend').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Date').font = header_font
        ws.cell(row=row, column=2, value='Sales').font = header_font
        ws.cell(row=row, column=3, value='Orders').font = header_font
        row += 1
        
        for item in data.get('sales_trend', []):
            ws.cell(row=row, column=1, value=item.get('date', ''))
            ws.cell(row=row, column=2, value=f"${item.get('sales', 0):.2f}")
            ws.cell(row=row, column=3, value=item.get('orders', 0))
            row += 1
        
        row += 2
        
        # Write top products section
        ws.cell(row=row, column=1, value='Top Products').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Product Name').font = header_font
        ws.cell(row=row, column=2, value='Units Sold').font = header_font
        ws.cell(row=row, column=3, value='Revenue').font = header_font
        row += 1
        
        for product in data.get('top_products', []):
            ws.cell(row=row, column=1, value=product.get('name', ''))
            ws.cell(row=row, column=2, value=product.get('sales', 0))
            ws.cell(row=row, column=3, value=f"${product.get('revenue', 0):.2f}")
            row += 1
        
        row += 2
        
        # Write order status distribution
        ws.cell(row=row, column=1, value='Order Status Distribution').font = title_font
        row += 1
        
        ws.cell(row=row, column=1, value='Status').font = header_font
        ws.cell(row=row, column=2, value='Count').font = header_font
        row += 1
        
        for status, count in data.get('order_status_distribution', {}).items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        filename = f"analytics_{export_type}_{timestamp}.xlsx"
        
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename

    async def get_recent_activity(self, limit: int = 100, since: datetime = None) -> List[dict]:
        """Get recent activity logs with user details and timestamps."""
        from models.activity_log import ActivityLog
        from sqlalchemy.orm import selectinload
        
        query = select(ActivityLog).options(
            selectinload(ActivityLog.user)
        ).order_by(ActivityLog.created_at.desc())
        
        if since:
            query = query.where(ActivityLog.created_at >= since)
        
        query = query.limit(limit)
        
        result = await self.db.execute(query)
        activities = result.scalars().all()
        
        return [
            {
                "id": str(activity.id),
                "user_id": str(activity.user_id) if activity.user_id else None,
                "user_name": activity.user.full_name if activity.user else "System",
                "action_type": activity.action_type,
                "description": activity.description,
                "metadata": activity.metadata,
                "created_at": activity.created_at.isoformat() if activity.created_at else None
            }
            for activity in activities
        ]
