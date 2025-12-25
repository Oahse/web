import csv
import io
import json
from typing import List, Dict, Any, Optional
from datetime import datetime, date
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pydantic import BaseModel

# Import Jinja2 and WeasyPrint
from jinja2 import Environment, FileSystemLoader, select_autoescape
from weasyprint import HTML, CSS


class ExportFilters(BaseModel):
    """Filters for export data"""
    start_date: Optional[date] = None
    end_date: Optional[date] = None
    customer_id: Optional[str] = None
    subscription_status: Optional[str] = None
    payment_status: Optional[str] = None
    variant_ids: Optional[List[str]] = None


class ExportResult(BaseModel):
    """Result of export operation"""
    content: bytes
    content_type: str
    filename: str
    format_type: str
    generated_at: datetime


class ExportService:
    """Consolidated service for exporting data to various formats"""
    
    def __init__(self):
        # Initialize Jinja2 environment for template rendering
        self.jinja_env = Environment(
            loader=FileSystemLoader("templates"),
            autoescape=select_autoescape(["html", "xml"])
        )
    
    @staticmethod
    def export_orders_to_csv(orders: List[Dict[str, Any]]) -> io.BytesIO:
        """Export orders to CSV format"""
        output = io.StringIO()
        
        if not orders:
            return io.BytesIO(b"No orders to export")
        
        # Define CSV headers
        headers = [
            'Order ID', 'Customer Name', 'Customer Email', 'Status', 
            'Payment Status', 'Total Amount', 'Items Count', 'Created At'
        ]
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for order in orders:
            user = order.get('user', {})
            writer.writerow({
                'Order ID': order.get('id', ''),
                'Customer Name': f"{user.get('firstname', '')} {user.get('lastname', '')}".strip(),
                'Customer Email': user.get('email', ''),
                'Status': order.get('status', ''),
                'Payment Status': order.get('payment_status', ''),
                'Total Amount': f"${order.get('total_amount', 0):.2f}",
                'Items Count': len(order.get('items', [])),
                'Created At': order.get('created_at', '')
            })
        
        # Convert StringIO to BytesIO
        bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        return bytes_output
    
    @staticmethod
    def export_orders_to_excel(orders: List[Dict[str, Any]]) -> io.BytesIO:
        """Export orders to Excel format"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"
        
        # Define headers
        headers = [
            'Order ID', 'Customer Name', 'Customer Email', 'Status', 
            'Payment Status', 'Total Amount', 'Items Count', 'Created At'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, order in enumerate(orders, 2):
            user = order.get('user', {})
            ws.cell(row=row_num, column=1, value=order.get('id', ''))
            ws.cell(row=row_num, column=2, value=f"{user.get('firstname', '')} {user.get('lastname', '')}".strip())
            ws.cell(row=row_num, column=3, value=user.get('email', ''))
            ws.cell(row=row_num, column=4, value=order.get('status', ''))
            ws.cell(row=row_num, column=5, value=order.get('payment_status', ''))
            ws.cell(row=row_num, column=6, value=order.get('total_amount', 0))
            ws.cell(row=row_num, column=7, value=len(order.get('items', [])))
            ws.cell(row=row_num, column=8, value=order.get('created_at', ''))
        
        # Adjust column widths
        column_widths = [15, 20, 30, 15, 15, 15, 12, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    def export_orders_to_pdf(orders: List[Dict[str, Any]]) -> io.BytesIO:
        """Export orders to PDF format using Jinja2 and WeasyPrint"""
        # Basic HTML template for the order export
        # In a real application, this would likely be loaded from a separate .html file
        template_string = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>Orders Export Report</title>
            <style>
                body { font-family: sans-serif; margin: 0; padding: 0; font-size: 10px; }
                .container { width: 100%; margin: 0 auto; padding: 20px; }
                h1 { text-align: center; color: #1a1a1a; font-size: 20px; margin-bottom: 20px; }
                .meta { text-align: center; margin-bottom: 30px; color: #555; }
                table { width: 100%; border-collapse: collapse; margin-bottom: 30px; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #4472C4; color: white; font-weight: bold; font-size: 11px;}
                tr:nth-child(even) { background-color: #f2f2f2; }
                .summary { margin-top: 20px; font-size: 11px; }
                .summary span { font-weight: bold; }
                .order-items { margin-top: 10px; border-top: 1px solid #eee; padding-top: 5px; font-size: 9px; }
                .order-item { margin-bottom: 3px; }
                .order-item-detail { margin-left: 15px; color: #666; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Orders Export Report</h1>
                <p class="meta">Generated on: {{ generation_date }}</p>

                {% if not orders %}
                    <p style="text-align: center;">No orders to export.</p>
                {% else %}
                    <div class="summary">
                        <span>Total Orders:</span> {{ orders|length }} |
                        <span>Total Revenue:</span> ${{ "%.2f"|format(total_revenue) }}
                    </div>

                    <table>
                        <thead>
                            <tr>
                                <th>Order ID</th>
                                <th>Customer</th>
                                <th>Status</th>
                                <th>Amount</th>
                                <th>Date</th>
                                <th>Items</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for order in orders %}
                            <tr>
                                <td>{{ order.id[:8] }}...</td>
                                <td>{{ order.customer_name }}</td>
                                <td>{{ order.status }}</td>
                                <td>${{ "%.2f"|format(order.total_amount) }}</td>
                                <td>{{ order.created_at[:10] }}</td>
                                <td>
                                    {% if order.items %}
                                        <div class="order-items">
                                            {% for item in order.items %}
                                                <div class="order-item">
                                                    {{ item.variant.product_name }} ({{ item.variant.name }}) - {{ item.quantity }} x ${{ "%.2f"|format(item.price_per_unit) }}
                                                    <div class="order-item-detail">Total: ${{ "%.2f"|format(item.total_price) }}</div>
                                                </div>
                                            {% endfor %}
                                        </div>
                                    {% else %}
                                        No items
                                    {% endif %}
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                {% endif %}
            </div>
        </body>
        </html>
        """

        # Setup Jinja2 environment to load from a string
        env = Environment(loader=FileSystemLoader("."), autoescape=select_autoescape(["html", "xml"]))
        template = env.from_string(template_string)

        total_revenue = sum(order.get('total_amount', 0) for order in orders)
        generation_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Prepare data for rendering
        rendered_orders = []
        for order in orders:
            user = order.get('user', {})
            customer_name = f"{user.get('firstname', '')} {user.get('lastname', '')}".strip() or 'N/A'
            rendered_orders.append({
                'id': order.get('id', ''),
                'customer_name': customer_name,
                'status': order.get('status', ''),
                'total_amount': order.get('total_amount', 0),
                'created_at': order.get('created_at', ''),
                'items': order.get('items', [])
            })

        html_content = template.render(
            orders=rendered_orders,
            total_revenue=total_revenue,
            generation_date=generation_date
        )

        # Convert HTML to PDF using WeasyPrint
        pdf_bytes = HTML(string=html_content).write_pdf()

        output = io.BytesIO(pdf_bytes)
        output.seek(0)
        return output
    
    # Enhanced export methods for subscriptions and payments
    
    async def export_subscription_data(
        self,
        subscriptions_data: List[Dict[str, Any]],
        filters: ExportFilters,
        format_type: str,
        template_name: Optional[str] = None
    ) -> ExportResult:
        """
        Export subscription data in various formats
        
        Args:
            subscriptions_data: List of subscription data dictionaries
            filters: Export filters applied
            format_type: Output format ("csv", "json", "html")
            template_name: Optional custom template name
            
        Returns:
            ExportResult with exported content
        """
        timestamp = datetime.now()
        
        if format_type == "csv":
            return await self._export_subscriptions_csv(subscriptions_data, filters, timestamp)
        elif format_type == "json":
            return await self._export_subscriptions_json(subscriptions_data, filters, timestamp)
        elif format_type == "html":
            template = template_name or "exports/subscriptions_export.html"
            return await self._export_subscriptions_html(subscriptions_data, filters, timestamp, template)
        else:
            raise ValueError(f"Unsupported format type: {format_type}")
    
    async def _export_subscriptions_csv(
        self,
        subscriptions_data: List[Dict[str, Any]],
        filters: ExportFilters,
        timestamp: datetime
    ) -> ExportResult:
        """Export subscriptions to CSV format"""
        output = io.StringIO()
        
        if not subscriptions_data:
            output.write("No subscription data to export\n")
        else:
            # Define CSV headers
            headers = [
                'Subscription ID', 'Customer Name', 'Customer Email', 'Status',
                'Total Cost', 'Currency', 'Delivery Type', 'Billing Cycle',
                'Variants', 'Admin Fee', 'Delivery Cost', 'Tax Amount',
                'Loyalty Discount', 'Created At', 'Next Billing Date'
            ]
            
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            
            for subscription in subscriptions_data:
                user = subscription.get('user', {})
                cost_breakdown = subscription.get('cost_breakdown', {})
                
                # Format variants list
                variants = subscription.get('variants', [])
                variant_names = [v.get('name', 'Unknown') for v in variants]
                
                writer.writerow({
                    'Subscription ID': subscription.get('id', ''),
                    'Customer Name': f"{user.get('firstname', '')} {user.get('lastname', '')}".strip(),
                    'Customer Email': user.get('email', ''),
                    'Status': subscription.get('status', ''),
                    'Total Cost': cost_breakdown.get('total_amount', 0),
                    'Currency': cost_breakdown.get('currency', 'USD'),
                    'Delivery Type': subscription.get('delivery_type', ''),
                    'Billing Cycle': subscription.get('billing_cycle', ''),
                    'Variants': '; '.join(variant_names),
                    'Admin Fee': cost_breakdown.get('admin_fee', 0),
                    'Delivery Cost': cost_breakdown.get('delivery_cost', 0),
                    'Tax Amount': cost_breakdown.get('tax_amount', 0),
                    'Loyalty Discount': cost_breakdown.get('loyalty_discount', 0),
                    'Created At': subscription.get('created_at', ''),
                    'Next Billing Date': subscription.get('next_billing_date', '')
                })
        
        content = output.getvalue().encode('utf-8')
        filename = f"subscriptions_export_{timestamp.strftime('%Y%m%d_%H%M%S')}.csv"
        
        return ExportResult(
            content=content,
            content_type='text/csv',
            filename=filename,
            format_type='csv',
            generated_at=timestamp
        )
    
    async def _export_subscriptions_json(
        self,
        subscriptions_data: List[Dict[str, Any]],
        filters: ExportFilters,
        timestamp: datetime
    ) -> ExportResult:
        """Export subscriptions to JSON format"""
        
        # Prepare export metadata
        export_data = {
            'metadata': {
                'generated_at': timestamp.isoformat(),
                'format': 'json',
                'filters_applied': filters.dict(exclude_none=True),
                'total_records': len(subscriptions_data)
            },
            'subscriptions': subscriptions_data
        }
        
        # Convert to JSON with proper serialization
        json_content = json.dumps(export_data, indent=2, default=self._json_serializer)
        content = json_content.encode('utf-8')
        filename = f"subscriptions_export_{timestamp.strftime('%Y%m%d_%H%M%S')}.json"
        
        return ExportResult(
            content=content,
            content_type='application/json',
            filename=filename,
            format_type='json',
            generated_at=timestamp
        )
    
    async def _export_subscriptions_html(
        self,
        subscriptions_data: List[Dict[str, Any]],
        filters: ExportFilters,
        timestamp: datetime,
        template_name: str
    ) -> ExportResult:
        """Export subscriptions to HTML format using Jinja template"""
        
        # Calculate summary statistics
        total_revenue = sum(
            sub.get('cost_breakdown', {}).get('total_amount', 0) 
            for sub in subscriptions_data
        )
        
        active_subscriptions = len([
            sub for sub in subscriptions_data 
            if sub.get('status') == 'active'
        ])
        
        context = {
            'subscriptions': subscriptions_data,
            'filters': filters.dict(exclude_none=True),
            'summary': {
                'total_subscriptions': len(subscriptions_data),
                'active_subscriptions': active_subscriptions,
                'total_revenue': total_revenue,
                'generated_at': timestamp
            },
            'company_name': 'Banwee'
        }
        
        try:
            template = self.jinja_env.get_template(template_name)
            rendered_content = template.render(context)
        except Exception:
            # Fallback to basic HTML if template not found
            rendered_content = self._generate_basic_html_export(subscriptions_data, "Subscriptions Export")
        
        content = rendered_content.encode('utf-8')
        filename = f"subscriptions_export_{timestamp.strftime('%Y%m%d_%H%M%S')}.html"
        
        return ExportResult(
            content=content,
            content_type='text/html',
            filename=filename,
            format_type='html',
            generated_at=timestamp
        )
    
    def _json_serializer(self, obj):
        """JSON serializer for special types"""
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        elif isinstance(obj, Decimal):
            return float(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    def _generate_basic_html_export(self, data: List[Dict[str, Any]], title: str) -> str:
        """Generate basic HTML export when template is not available"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>{title}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p>Total records: {len(data)}</p>
            <table>
                <tr><th>Data</th></tr>
        """
        
        for item in data:
            html += f"<tr><td>{str(item)}</td></tr>"
        
        html += """
            </table>
        </body>
        </html>
        """
        return html