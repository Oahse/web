"""
Property-based tests for analytics export functionality.

**Feature: app-enhancements, Properties 17-19: Analytics export**
**Validates: Requirements 6.1, 6.2, 6.3**
"""

import pytest
from hypothesis import given, strategies as st, settings, HealthCheck
from datetime import datetime, timedelta
from sqlalchemy.ext.asyncio import AsyncSession
from services.analytics import AnalyticsService
import csv
import io


# Strategy for generating analytics data
@st.composite
def analytics_data_strategy(draw):
    """Generate random analytics data for testing."""
    return {
        'total_sales': draw(st.floats(min_value=0, max_value=1000000, allow_nan=False, allow_infinity=False)),
        'total_orders': draw(st.integers(min_value=0, max_value=10000)),
        'total_users': draw(st.integers(min_value=0, max_value=10000)),
        'total_products': draw(st.integers(min_value=0, max_value=1000)),
        'conversion_rate': draw(st.floats(min_value=0, max_value=100, allow_nan=False, allow_infinity=False)),
        'average_order_value': draw(st.floats(min_value=0, max_value=10000, allow_nan=False, allow_infinity=False)),
        'sales_trend': [
            {
                'date': (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d'),
                'sales': draw(st.floats(min_value=0, max_value=10000, allow_nan=False, allow_infinity=False)),
                'orders': draw(st.integers(min_value=0, max_value=100))
            }
            for i in range(draw(st.integers(min_value=1, max_value=30)))
        ],
        'top_products': [
            {
                'name': f"Product {i}",
                'sales': draw(st.integers(min_value=0, max_value=1000)),
                'revenue': draw(st.floats(min_value=0, max_value=50000, allow_nan=False, allow_infinity=False))
            }
            for i in range(draw(st.integers(min_value=0, max_value=10)))
        ],
        'order_status_distribution': {
            status: draw(st.integers(min_value=0, max_value=1000))
            for status in draw(st.lists(
                st.sampled_from(['pending', 'processing', 'confirmed', 'shipped', 'delivered', 'cancelled']),
                min_size=1,
                max_size=6,
                unique=True
            ))
        }
    }


@pytest.mark.asyncio
@given(data=analytics_data_strategy())
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
async def test_property_export_data_accuracy(data, db_session: AsyncSession):
    """
    Property 17: Analytics export data accuracy
    
    For any export operation, the generated file should contain the current 
    analytics data matching applied filters.
    
    **Validates: Requirements 6.1**
    """
    # Create analytics service
    analytics_service = AnalyticsService(db_session)
    
    # Export as CSV
    csv_content, csv_content_type, csv_filename = await analytics_service.export_data(
        data=data,
        format='csv',
        export_type='dashboard'
    )
    
    # Verify CSV content contains the data
    csv_text = csv_content.decode('utf-8')
    
    # Check that key metrics are present in the CSV
    assert f"{data['total_sales']:.2f}" in csv_text or str(int(data['total_sales'])) in csv_text, \
        "Total sales should be in CSV export"
    assert str(data['total_orders']) in csv_text, \
        "Total orders should be in CSV export"
    assert str(data['total_users']) in csv_text, \
        "Total users should be in CSV export"
    assert str(data['total_products']) in csv_text, \
        "Total products should be in CSV export"
    
    # Check that sales trend data is present
    for trend_item in data['sales_trend'][:3]:  # Check first 3 items
        assert trend_item['date'] in csv_text, \
            f"Sales trend date {trend_item['date']} should be in CSV export"
    
    # Check that top products are present
    for product in data['top_products'][:3]:  # Check first 3 products
        assert product['name'] in csv_text, \
            f"Product {product['name']} should be in CSV export"
    
    # Verify content type and filename
    assert csv_content_type == 'text/csv', \
        "CSV export should have correct content type"
    assert csv_filename.endswith('.csv'), \
        "CSV export should have .csv extension"


@pytest.mark.asyncio
@given(data=analytics_data_strategy())
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
async def test_property_csv_format_correctness(data, db_session: AsyncSession):
    """
    Property 18: CSV export format correctness
    
    For any CSV export, the file should be properly formatted and parseable 
    as valid CSV.
    
    **Validates: Requirements 6.2**
    """
    # Create analytics service
    analytics_service = AnalyticsService(db_session)
    
    # Export as CSV
    csv_content, content_type, filename = await analytics_service.export_data(
        data=data,
        format='csv',
        export_type='dashboard'
    )
    
    # Verify it's valid CSV by parsing it
    csv_text = csv_content.decode('utf-8')
    csv_reader = csv.reader(io.StringIO(csv_text))
    
    rows = list(csv_reader)
    
    # Verify we have rows
    assert len(rows) > 0, "CSV should contain rows"
    
    # Verify structure - should have sections
    # Find the summary section
    summary_found = False
    sales_trend_found = False
    top_products_found = False
    
    for row in rows:
        if row and 'Analytics Summary' in row[0]:
            summary_found = True
        if row and 'Sales Trend' in row[0]:
            sales_trend_found = True
        if row and 'Top Products' in row[0]:
            top_products_found = True
    
    assert summary_found, "CSV should contain Analytics Summary section"
    assert sales_trend_found, "CSV should contain Sales Trend section"
    assert top_products_found, "CSV should contain Top Products section"
    
    # Verify no parsing errors occurred
    for i, row in enumerate(rows):
        assert isinstance(row, list), f"Row {i} should be a list"
    
    # Verify content type
    assert content_type == 'text/csv', "Content type should be text/csv"
    
    # Verify filename format
    assert filename.startswith('analytics_'), "Filename should start with 'analytics_'"
    assert filename.endswith('.csv'), "Filename should end with '.csv'"


@pytest.mark.asyncio
@given(data=analytics_data_strategy())
@settings(
    max_examples=100,
    deadline=None,
    suppress_health_check=[HealthCheck.function_scoped_fixture]
)
async def test_property_excel_format_correctness(data, db_session: AsyncSession):
    """
    Property 19: Excel export format correctness
    
    For any Excel export, the file should be properly formatted and openable 
    as valid XLSX.
    
    **Validates: Requirements 6.3**
    """
    # Create analytics service
    analytics_service = AnalyticsService(db_session)
    
    # Export as Excel
    xlsx_content, content_type, filename = await analytics_service.export_data(
        data=data,
        format='xlsx',
        export_type='dashboard'
    )
    
    # Verify it's valid Excel by opening it
    try:
        from openpyxl import load_workbook
    except ImportError:
        pytest.skip("openpyxl not installed")
    
    # Load the workbook from bytes
    wb = load_workbook(io.BytesIO(xlsx_content))
    
    # Verify workbook structure
    assert len(wb.worksheets) > 0, "Excel file should contain at least one worksheet"
    
    ws = wb.active
    assert ws.title == "Analytics Summary", "First worksheet should be named 'Analytics Summary'"
    
    # Verify the worksheet has data
    assert ws.max_row > 0, "Worksheet should contain rows"
    assert ws.max_column > 0, "Worksheet should contain columns"
    
    # Verify key sections exist in the worksheet
    cell_values = []
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), values_only=True):
        cell_values.extend([str(cell) for cell in row if cell is not None])
    
    content_text = ' '.join(cell_values)
    
    assert 'Analytics Summary' in content_text, "Excel should contain Analytics Summary section"
    assert 'Sales Trend' in content_text, "Excel should contain Sales Trend section"
    assert 'Top Products' in content_text, "Excel should contain Top Products section"
    
    # Verify content type
    assert content_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', \
        "Content type should be correct for Excel files"
    
    # Verify filename format
    assert filename.startswith('analytics_'), "Filename should start with 'analytics_'"
    assert filename.endswith('.xlsx'), "Filename should end with '.xlsx'"
    
    # Close workbook
    wb.close()
